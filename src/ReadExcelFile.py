import datetime
import numbers

import openpyxl

#Create a new workbook
workbook = openpyxl.load_workbook("Superstore.xlsx")

#Select the active sheet
worksheet = workbook.active

#Read the rows from active sheet
#for row in worksheet.iter_rows(values_only=True):
    #print(row)

#Get Unique products count
unique_products = set()
for column_data in worksheet['Q']:
    if column_data.value not in unique_products:
        unique_products.add(column_data.value)

print(f"There are {len(unique_products)} unique products.")

#Get Total customers
Total_Customers = set()
for column_data in worksheet['G']:
    if column_data.value not in Total_Customers:
        Total_Customers.add(column_data.value)
print(f"There are {len(Total_Customers)} total customers.")
print(Total_Customers)

#Unique years
years_count = set()
for column_data in worksheet['C']:
    if column_data.value not in years_count:
        if isinstance(column_data.value, datetime.date):
            year = column_data.value.year
            years_count.add(year)
print(years_count)

#Get Total Sales amount
total_sales_amount = 0
for column_data in worksheet['R']:
    if isinstance(column_data.value, int):
        total_sales_amount += column_data.value
print(total_sales_amount)

#Get Total profit
total_profit = 0
total_loss = 0
for column_data in worksheet['U']:
    if isinstance(column_data.value, numbers.Number):
        if column_data.value > 0:
            total_profit += column_data.value
        elif column_data.value < 0:
            total_loss += column_data.value

print("Profit in positive: ", total_profit)
print("Loss in negative: ", total_loss)
actual_sales_profit = round(total_profit, 2) + round(total_loss, 2)
print("Actual sales profit: ", actual_sales_profit)

#Get Quantity count for all the years
total_quantity = 0
for column_data in worksheet['S']:
    if isinstance(column_data.value, str):
        print("do nothing")
    elif isinstance(column_data.value, numbers.Number):
        total_quantity += column_data.value
print("Total quantity: ", total_quantity)

#Get Quantity per year
header = [cell.value for cell in worksheet[1]]
year_col_idx = header.index('Order Date')+1
qty_col_idx = header.index('Quantity')+1

year_wise_quantity = {}
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skip header row
    year = row[year_col_idx - 1]
    year = year.year
    qty = row[qty_col_idx - 1]
    if year is not None and qty is not None:
        year_wise_quantity[year] = year_wise_quantity.get(year, 0) + qty
# Display results
for year, total_qty in sorted(year_wise_quantity.items()):
    print(f"Year: {year}, Total Quantity: {total_qty}")
print(year_wise_quantity)
print(dict(sorted(year_wise_quantity.items())))

#Get year wise sales
sales_col_idx = header.index('Sales')+1
year_wise_sales = {}
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    sales = row[sales_col_idx - 1]
    if year is not None and sales is not None:
        year_wise_sales[year] = year_wise_sales.get(year, 0) + round(sales, 2)
for year, sales_qty in sorted(year_wise_sales.items()):
    print(f"Year: {year}, Sales Quantity: {round(sales_qty, 2)}")

#Get year wise profit and loss
year_wise_earning ={}
profit_col_idx= header.index('Profit')+1
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    earning = row[profit_col_idx -1]
    if year is not None and earning is not None:
        year_wise_earning[year] = year_wise_earning.get(year, 0) + earning
for year, profit in sorted(year_wise_earning.items()):
    print(f"Year: {year}, Profit: {round(profit, 2)}")


#year wise ship modes
# Read header to get column indices
ship_mode_col_idx = header.index('Ship Mode') + 1

# Prepare data structure: {year: {ship_mode: count}}
ship_mode_summary = {}

for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    ship_mode = row[ship_mode_col_idx - 1]
    if year and ship_mode:
        if year not in ship_mode_summary:
            ship_mode_summary[year] = {}
        ship_mode_summary[year][ship_mode] = ship_mode_summary[year].get(ship_mode, 0) + 1
print(ship_mode_summary)
# Print result
for year in sorted(ship_mode_summary):
    print(f"Year: {year}")
    for mode, count in ship_mode_summary[year].items():
        print(f"   Ship Mode: {mode} - {count} orders")

# Prepare the nested summary dictionary
# Format: summary[year][state] = {'Sales': ..., 'Profit': ..., 'Quantity': ...}
summary = {}
state_col_idx = header.index('State')+1
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skips header
    year = row[year_col_idx -1]
    year = year.year
    state = row[state_col_idx -1]
    sales = row[sales_col_idx -1]
    profit = row[profit_col_idx -1]
    qty = row[qty_col_idx -1]

    if year and state:
        if year not in summary:
            summary[year] = {}
        if state not in summary[year]:
            summary[year][state] = {'Sales': 0, 'Profit': 0, 'Quantity': 0}
        summary[year][state]['Sales'] += float(sales)
        summary[year][state]['Profit'] += float(profit)
        summary[year][state]['Quantity'] += qty

# Display the summarized data
for year in sorted(summary):
    print(f"\nYear: {year}")
    for state, data in summary[year].items():
        print(f"  State: {state} | Sales: {data['Sales']} | Profit: {data['Profit']} | Quantity: {data['Quantity']}")


#Get year wise Discount offered
year_wise_discount ={}
discount_col_idx= header.index('Discount')+1
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    discount = row[discount_col_idx -1]
    if year is not None and discount is not None:
        year_wise_discount[year] = year_wise_discount.get(year, 0) + discount
for year, discount in sorted(year_wise_discount.items()):
    print(f"Year: {year}, Discount: {round(discount, 2)}")
