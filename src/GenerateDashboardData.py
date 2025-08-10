from flask import Flask, render_template
import openpyxl
import datetime
import numbers



#Create a new workbook
workbook = openpyxl.load_workbook("Superstore.xlsx")

#Select the active sheet
worksheet = workbook.active

#Get Unique products count
unique_products = set()
for column_data in worksheet['Q']:
    if column_data.value not in unique_products:
        unique_products.add(column_data.value)
Total_product_count = len(unique_products)

#Get Total customers
Total_Customers = set()
for column_data in worksheet['G']:
    if column_data.value not in Total_Customers:
        Total_Customers.add(column_data.value)
Total_Customers_Count = len(Total_Customers)

#Unique years
years_count = set()
for column_data in worksheet['C']:
    if column_data.value not in years_count:
        if isinstance(column_data.value, datetime.date):
            year = column_data.value.year
            years_count.add(year)
years_count = sorted(years_count)

#Get Total profit
total_profit = 0
total_loss = 0
for column_data in worksheet['U']:
    if isinstance(column_data.value, numbers.Number):
        if column_data.value > 0:
            total_profit += column_data.value
        elif column_data.value < 0:
            total_loss += column_data.value
actual_sales_profit = round(total_profit, 2) + round(total_loss, 2)

#Get Quantity per year
header = [cell.value for cell in worksheet[1]]
year_col_idx = header.index('Order Date')+1
qty_col_idx = header.index('Quantity')+1
year_wise_prod_countity = {}
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skip header row
    year = row[year_col_idx - 1]
    year = year.year
    qty = row[qty_col_idx - 1]
    if year is not None and qty is not None:
        year_wise_prod_countity[year] = year_wise_prod_countity.get(year, 0) + qty
dict(sorted(year_wise_prod_countity.items()))


#Get year wise sales
sales_col_idx = header.index('Sales')+1
year_wise_sales = {}
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    sales = row[sales_col_idx - 1]
    sales = round(sales, 2)
    if year is not None and sales is not None:
        year_wise_sales[year] = year_wise_sales.get(year, 0) + sales
dict(sorted(year_wise_sales.items()))

#Get year wise profit and loss
year_wise_earning ={}
profit_col_idx= header.index('Profit')+1
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    earning = row[profit_col_idx -1]
    if year is not None and earning is not None:
        year_wise_earning[year] = year_wise_earning.get(year, 0) + earning
dict(sorted(year_wise_earning.items()))

#year wise ship modes
ship_mode_col_idx = header.index('Ship Mode') + 1
# Prepare data structure: {year: {ship_mode: count}}
ship_mode_summary = {}
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    ship_mode = row[ship_mode_col_idx - 1]
    if year and ship_mode:
        if year not in ship_mode_summary:
            ship_mode_summary[year] = {}
        ship_mode_summary[year][ship_mode] = ship_mode_summary[year].get(ship_mode, 0) + 1

#State wise yearly data
# Prepare the nested summary dictionary
# Format: summary[year][state] = {'Sales': ..., 'Profit': ..., 'Quantity': ...}
summary = {}
state_col_idx = header.index('State')+1
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skips header
    year = row[year_col_idx -1]
    year = year.year
    state = row[state_col_idx -1]
    sales = row[sales_col_idx -1]
    profit = row[profit_col_idx -1]
    qty = row[qty_col_idx -1]

    if year and state:
        if year not in summary:
            summary[year] = {}
        if state not in summary[year]:
            summary[year][state] = {'Sales': 0, 'Profit': 0, 'Quantity': 0}
        summary[year][state]['Sales'] += float(sales)
        summary[year][state]['Profit'] += float(profit)
        summary[year][state]['Quantity'] += qty

#Get year wise Discount offered
year_wise_discount ={}
discount_col_idx= header.index('Discount')+1
for row in worksheet.iter_rows(min_row=2, values_only=True):
    year = row[year_col_idx - 1]
    year = year.year
    discount = row[discount_col_idx -1]
    if year is not None and discount is not None:
        year_wise_discount[year] = year_wise_discount.get(year, 0) + discount

#Pass the data to HTML to display to end user
app = Flask(__name__, template_folder='../templates')
@app.route('/')
def index():
    # Sample data to pass to the HTML template
    user_data = {"name": "Alice", "age": 30}
    return render_template('SalesDashboard.html', productCount=Total_product_count, customerCount=Total_Customers_Count, years=years_count,
                           profit=actual_sales_profit, quantity=year_wise_prod_countity, sales_yearly=year_wise_sales, earning_yearly=year_wise_earning,
                           ship_mode=ship_mode_summary, yearly_discount=year_wise_discount, summary=summary)

if __name__ == '__main__':
    app.run(debug=True)

